matrix = [[0, 1],  [2, 3]]
letters = ['a', 'b', 'c', 'd', 'e']
zeros = [0] * 5
combined = zeros + letters

numbers = list(range(10))   # 0 - 9
step_numbers = numbers[::3] # 0 3 6 9
reversed_numbers = numbers[::-1] # 9 - 0
even_numbers = list(range(2, 11, 2)) # 2 4 6 8 10
odd_numbers = list(range(1, 11, 2))  # 1 3 5 7 9

chars = list('Hello World')
chars2 = chars[:]
chars3 = chars[::3]

# unpacking list
a, b, *other, last = numbers

# for letter in letters:
#   print(letter)

# for index, letter in enumerate(letters):
#   print(index, letter)

letters.append('f')
letters.insert(2, 'xyz')
letters.pop(1)
letters.index('e')
letters.count('xyz')
letters.remove('d')
letters.reverse()
letters.clear()
del letters[0:3]

mess_numbers = [3, 51, 2, 8, 6]
mess_numbers.sort(reverse=True)
ordered_numbers = sorted(mess_numbers, reverse=True)

# lambda function 匿名函数排序
def sort_item(item):
  return item[1]
items = [('product3', 30), ('product1', 10), ('product2', 5),]
items.sort(key=sort_item)
items.sort(key=lambda item: item[1])

# prices = []
# for item in items:
#   prices.append(item[1])
# print(prices)
# 匿名函数在map跟filter方法里的运用
prices = list(map(lambda item: item[1], items))
f_prices = list(filter(lambda item: item[1] >= 10, items))
# List Comprehension 代替map方法
# [expression for item in items]
prices_list = [item[1] for item in items] # map
filter_prices = [item for item in items if item[1] >= 10] # filter


# Zip
list1 = [1, 2, 3]
list2 = [10, 20, 30]
list3 = list(zip(list1, list2))
list4 = list(zip(list1, list2, 'abc'))


# stack
stack = []
stack.append(1)
stack.append(2)
stack.append(3)
stack.pop()
if not stack:
  print('empty stack')


# queue
from collections import deque
queue = deque([1 ,2, 3])
queue.appendleft('a')
queue.popleft()


#tuple
point0 = (1, 2)
point1 = tuple([1, 2])
point2 = tuple('hello')
point3 = (2, 4) * 3
point4 = (1, 3, 5)
a, b, c = point4
if 10 in point4:
  print('exists')


# swapping variable (背后道理是利用tuple)
x = 10
y = 20
x , y = y, x    # 同等x, y = (y, x) 也就是x, y = (20, 10), unpacking


# array - 处理大量数字 性能优化
from array import array
nums_array = array('i', [1, 2, 3])  # 'i'第一个参数指定array数据类型
nums = list(nums_array)


# set数据类型: un-ordered collection of Unique items, no index所以不能nums[0]读取值
dup_nums = [1, 1, 2, 3, 4, 4]
uniques = set(dup_nums)   # {1, 2, 3, 4}
uniques_2 = {1, 5}
union = uniques | uniques_2 # union {1,2,3,4,5}
intersect = uniques & uniques_2 # {1}
difference = uniques - uniques_2  # {2, 3, 4}
not_both = uniques ^ uniques_2 # {2,3,4,5}


# dictionary: key - value pair
point = {'x': 1, 'y': 2, 'z': 3, 'w': 4}
point2 = dict(a=3, b=4)
del point['w']
# 下面两个for 返回一样结果
for key in point:
  print(key, point[key])
for key, value in point.items():
  print(key, value)


# Dictionary Comprehensions
values = [item * 2 for item in range(5)] # list [0,2,4,6,8]
sets = {item * 2 for item in range(5)} # sets {0,2,4,6,8}
dics = {item: item * 2 for item in range(5)} # dics {0: 0, 1: 2, 2: 4, 3: 6, 4: 8}
tuples = (item * 2 for item in range(5)) # generator


# Generator object: iterable 用来处理大量数据，只有在循环处理它时才会生成，所以省size
values = (item * 2 for item in range(1000000)) 
print(type(values)) # <class 'generator'>
for x in values:
  print(x)


# unpacking operator * 跟js里的spread ...三个点是一样的
# print(*numbers)  # 1 2 3 4 5
# 用两个** 来unpack dictionary
unpacked_nums = [*range(5)]  # 等于 list(range(5))
unpacked_strs = [*'hello world']
first_dic = {'x': 1, 'y': 2}
second_dic = {'z': 3}
third_dic = {**first_dic, **second_dic, 'yy': 44}


# Exercise
sentence = 'This is a common interview question'
char_frequency = {}
for char in sentence:
  if char in char_frequency:
    char_frequency[char] += 1
  else:
    char_frequency[char] = 1
result = sorted(char_frequency.items(), key=lambda kv: kv[1], reverse=True)
# print(result[0])  # {'i': 5}
  
# Exception---------------------------------------------
try:
  file = open('app.py')
  value = int(input('age:'))
  xfactor = 10 / value
  # file.close() # 如果前面出错，这行不执行，文件就没关闭,但是下面用with打开文件自动关闭
  with open('app.py') as source, open('test.txt') as target:
    print('File opened.')
except (ValueError, ZeroDivisionError) as err:
  print('your input is invaild!')
else:
  print(value)
finally:
  file.close() # 释放外部资源：如关闭文件，断开数据库连接等操作，网路连接

# raise exception 不推荐这么做的cost 
def calc_age(age):
  if age <= 0:
    raise ValueError('Age cannot be Zero or less.')
  return 10 / age

try:
  calc_age(-1)
except ValueError as error:
  # print(error)
  pass

# Class---------------------------------------------
# Class: blueprint for creating new objects (Human)
# Object: an instance of a class (Tom, Emma, Jimmy)
# class bundles data (age, gender) and functions (run, eat) into one unit
# isinstance(point, Point)
# class Name can't have underscore '_' -> MyBigPoint
class Point:
  # class level attribute shared to all instances
  default_color = 'red' 
  
  # instance level attribute only applys to that one instance
  def __init__(self, x, y):
    self.x = x
    self.y = y
    
  # magic method
  def __str__(self):  
    return f'({self.x}, {self.y})'
  def __eq__(self, other):
    return self.x == other.x and self.y == other.y
  def __gt__(self, other):
    return self.x > other.x and self.y > other.y
  def __add__(self, other):
    # return f'({self.x + other.x}, {self.y + other.y})'
    return Point(self.x + other.x, self.y + other.y)

  # class level method / factory method
  @classmethod
  def zero(cls):
    return cls(0, 0) # cls(0,0) == Point(0,0)

  # instance level method
  def draw(self):
    print(f'Point ({self.x}, {self.y})')

point_0 = Point.zero()
point_1 = Point(10, 20)
point_2 = Point(1, 2)
point_3 = point_1 + point_2

# create custom container
class TagCloud:
  def __init__(self):
    self.__tags = {}    # __tags两个下划线表示私有属性
  
  def add(self, tag):
    self.__tags[tag.lower()] = self.__tags.get(tag, 0) + 1

  def __getitem__(self, tag):
    return self.__tags.get(tag.lower(), 0)
  
  def __setitem__(self, tag, count):
    self.__tags[tag.lower()] = count

  def __len__(self):
    return len(self.__tags)
  
  def __iter__(self):
    return iter(self.__tags)

cloud = TagCloud()
cloud.add('Python')
cloud.add('python')
cloud.add('python')
cloud.add('java')

# properties
class Product:
  def __init__(self, price):
    self.price = price

  @property
  def price(self):
    return self.__price
  
  @price.setter
  def price(self, value):
    if value < 0:
      raise ValueError('Price cannot be negative.')
    self.__price = value

product = Product(50)
product.price = 33
print(product.price)


# 继承
class Animal:
  def __init__(self):
    self.age = 10

  def eat(self):
    print('eat')
  
class Mammal(Animal):
  def __init__(self):
    super().__init__()
    self.weight = 2

  def walk(self):
    print('walk')

class Fish(Mammal):
  def swim(self):
    print('swim')

animal = Animal()
mammal = Mammal()
fish = Fish()

# 抽象类
from abc import ABC, abstractmethod
class Shape(ABC):
  def __init__(self, color):
    self.color = color
  @abstractmethod
  def calc_size(self):
    pass
  
  def show_color(self):
    return f'My color is {self.color}'
  @classmethod
  def zero(cls):
    return cls('white')
  
class Circle(Shape):
  def __init__(self, radius):
    super().__init__('red')
    self.radius = radius
  def calc_size(self):
    return f'Circle size is {3.14 * self.radius (**) 2}'
  
class Square(Shape):
  def __init__(self, length):
    super().__init__('green')
    self.length = length
  def calc_size(self):
    return f'Square size is {self.length **2 }'
  
shape = Shape('blue')   # 抽象类不能实例化, 会报错
circle = Circle(2)
square = Square(2)
# polymorphism 多态性 many forms
def calc(shapes):
  for shape in shapes:    
    print(shape.calc_size())
calc([circle, square])   

# 抽象类不能实例化，只有它的subclass可以, 并且子类必须实现父类的抽象方法
# 这样的好处是 所有子类都必须遵循统一的接口你，等于是一个合同contract
# abc - abstract base class
from abc import ABC, abstractmethod
class InvalidOperationError(Exception):
  pass

class Stream(ABC):
  def __init__(self):
    self.opened = False
  def open(self):
    if self.opened:
      raise InvalidOperationError('Stream is already opened.')
    self.open = True
  def close(self):
    if not self.opened:
      raise InvalidOperationError('Stream is already closed.')
    self.open = False
  @abstractmethod   # 抽象方法 - 子类必须implement它
  def read(self):
    pass

class FileStream(Stream):
  def read(self):
    print('Reading data from a file.')
class NetworkStream(Stream):
  def read(self):
    print('Reading data from network stream')
class MemoryStream(Stream):
  def read(self):
    print('Reading data from memory stream.')
file = FileStream()
network = NetworkStream()
memory = MemoryStream()
def get_data(streams):
  for stream in streams:
    stream.read()
get_data([file, network, memory])

# Extend build-in Types
class Text(str):
  def duplicate(self):
    return self + self

text = Text('Python')
# print(text.duplicate())

class TrackList(list):
  def append(self, obj):
    print('Track called.')
    super().append(obj)

list = TrackList()
# list.append(111)
# list.append(222)
# print(list)


# Data Classes
from collections import namedtuple
Position = namedtuple('Pos', ['x', 'y', 'z'])
pos1 = Position(x=1,y=2,z=3)
pos2 = Position(1,2,3)
# print(pos1 == pos2)


# ----------08-Modules
# import sys
# print(sys.path)

from ecommerce.shopping.sales import calc_shipping, calc_tax
calc_shipping()
calc_tax()

from ecommerce.shopping import sales
sales.calc_shipping()
sales.calc_tax()
# dir
# print(dir(sales))
# print(sales.__name__)
# print(sales.__package__)
# print(sales.__file__)



# ----------09_01-02 标准库 Python Standard Library
from pathlib import Path

path = Path('ecommerce/__init__.py')  # relative path
path = Path() / Path('ecommerce') / '__init__.py'  # connected path

print(path)           # ecommerce/__init__.py
print(path.name)      # __init__.py
print(path.stem)      # __init__
print(path.suffix)    # .py
print(path.parent)    # ecommerce
print(path.exists())  # True
print(path.is_file()) # True
print(path.is_dir())  # False
print(Path.home())    # /Users/shao
print(path.absolute())# /Users/shao/Desktop/pyGo/ecommerce/__init__.py
print(path.stat())    # detail info of the file
path.read_text()
path.read_bytes()
path.write_text('balabala')
path.write_bytes()
# rename the path (not file)
path = path.with_name('file.txt')
path = path.with_suffix('.json')
path = path.with_stem('test')
# folders and files
path = Path('ecommerce')
directories = [p for p in path.iterdir() if p.is_dir()] # all directories
files = [p for p in path.rglob('*.py')]                 # all python files
# 复制文件
import shutil
source = Path('ecommerce/__init__.py')
target = Path() / '__init__.py'
shutil.copy(source, target)


# ----------09_05 Zip files 打包
# 写入：打包压缩ecommerce下的所有目录和文件
from zipfile import ZipFile
from pathlib import Path

with ZipFile('files.zip', 'w') as zip:
  for path in Path('ecommerce').rglob('*.*'):
    zip.write(path)
# 读取：解压缩
with ZipFile('files.zip') as zip:
  print(zip.namelist())
  info = zip.getinfo('ecommerce/__init__.py')
  print(info.compress_size)
  print(info.file_size)
  zip.extractall('zipped')


# ----------09_06 操作CSV文件
import csv
# file = open('data.csv', 'w')
# file.close()    # 老办法不用with，每次要close

with open('data.csv', 'w') as file:
  writer = csv.writer(file)
  writer.writerow(['transaction id', 'product_id', 'price'])
  writer.writerow([1000, 1, 5])
  writer.writerow([1001, 2, 15])
  writer.writerows([[2000, 3, 22], [2002, 4, 33]])

with open('data.csv') as file:
  reader = csv.reader(file)
  # print(list(reader))
  for row in reader:
    print(row)


# ----------09_07 操作json文件
import json
from pathlib import Path
# --- write to json file ---
movies = [
  {'id': 1, 'title': 'Water World', 'year': 1990},
  {'id': 2, 'title': 'Giant Bike', 'year': 2000},
]
data = json.dumps(movies)
Path('movies.json').write_text(data)
# --- read from json file ---
data = Path('movies.json').read_text()  # string
movies = json.loads(data) # list [{'id':1, 'title': 'SkyFall'}, ...]


# ----------09_08 SQLite Database
import json
from pathlib import Path
import sqlite3

movies = json.loads(Path('movies.json').read_text())
with sqlite3.connect('db.sqlite3') as conn:
    command = 'INSERT INTO Movies VALUES(?, ?, ?)'
    for movie in movies:
        conn.execute(command, tuple(movie.values()))
    conn.execute(command, tuple([3, 'NeZha2', 2025]))
    conn.commit()

    command = 'SELECT * FROM Movies'
    result = conn.execute(command)
    for row in result:
        print(row)
    movie_list = result.fetchall()
    print(movie_list)


# ----------09_09-11 time / datetime moduels
import time
from datetime import datetime, timedelta
# create datetime object
seconds = time.time()
current = datetime.now()
dt0 = datetime.fromtimestamp(seconds)
dt1 = datetime(2025, 2, 28, 15, 30, 25)
dt2 = datetime.strptime('2025/01/02', '%Y/%m/%d') # 把字符串变dt object
dt_str1 = dt2.strftime('%Y/%m') # 把dt object 变成字符串
dt_str2 = f'{dt2.year}/{dt2.month}'
# timedelta 时间增量
dt3 = dt0 + timedelta(days=1, hours=3, seconds=10, minutes=5) 

duration = dt1 - dt2
print(duration, duration.total_seconds())


# ----------09_12-random value moduel随机值生成
import random
import string
random.random(), 
random.randint(1, 10), 
random.choice([*range(8)]),
random.choices([1, 2, 3, 4, 5], k=2),
''.join(random.choices(string.ascii_letters + string.digits, k=10))

nums = [1, 2, 3, 4, 5]
print(nums)
random.shuffle(nums)
print(nums)


# ----------09_13-open browser
import webbrowser
print('deployment done, open browser')
webbrowser.open('https://www.baidu.com')


# ----------09_14_15 - send email
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
import smtplib
from string import Template
from pathlib import Path

template = Template(Path('template.html').read_text())
message = MIMEMultipart()
message['from'] = 'bright shao'
message['to'] = 'nzsyl0203@gmail.com'
message['subject'] = 'this is a test mail'
email_body = template.substitute({'name': 'John', 'price': 1000})
message.attach(MIMEText(email_body, 'html'))
message.attach(MIMEImage(Path('test.png').read_bytes()))

with smtplib.SMTP(host='smtp.gmail.com', port=587) as smtp:
    smtp.ehlo()
    smtp.starttls()
    smtp.login('nzsyl0203@gmail.com', 'Syl0203!')
    smtp.send_message(message)
    print('Sent...')


# ----------09_16-command line arguments
# 命令行执行 python3 app.py -a -b -c, # print(sys.argv) 得到['app.py','-a','-b','-c']
import sys

if len(sys.argv) == 1:
  print('USAGE: python3 app.py <password>')
else:
  print(f'PASSWORD: {sys.argv[1]}')


# ----------09_17 - Running external programs
import subprocess
# 执行外部python文件
subprocess.run(['python3', 'other.py'])
# 执行 ls -l 命令
result = subprocess.run(['ls', '-l'],
                        capture_output=True,  # save output into result.stdout
                        text=True,)
print(result.args, result.returncode, result.stderr, result.stdout)
# 执行遇到错误
try:
    false_script = subprocess.run(['false'], check=True)
except subprocess.CalledProcessError as ex:
    print(ex)



# ----------10_1 & 2 Pypi Python Package Index
# 安装指定version的package使用双等号==，或者~=
pip3 install requests==2.9.* # 跟 ~=2.9.0 是一样的
pip3 install requests==2.*

import requests
response = requests.get('https://www.baidu.com')
print(response.status_code)

# ----------10_3 Virtual Environment
python3 -m venv env
source env/bin/activate

# ----------10_4 Pipenv，命令行下面：
pip3 install pipenv
pipenv install requests
pipenv --venv
pip3 uninstall requests
python3 app.py # 报错，找不到requests module，用下面的命令激活env
pipenv shell
python3 app.py # 不报错了 能找到requests module，如果想deactivate，用exit
exit

# ----------10_8 Publishing packages
# 先安装3个包稍后用: pip3 install setuptools wheel twine
# 然后创建一个project，包含文件brightpdf/shaohi/data/tests/LICENSE/README.md/setup.py
# 打包: python3 setup.py sdist bdist_wheel
# 发布: twine upload dist/*  (不好用，用右边命令) python3 -m twine upload dist/*
# 下载使用： pip3 install brightpypi
from brightpypi import pdf2text, pdf2image, hello, guess

converter = pdf2text.Converter()
converter.convert('abc')
print(pdf2text.hahadaxiao())
pdf2image.convert()
hello.sayhi('John')
guess.guess(100)

# ----------10_09 Docstrings 即""" module的介绍包括class和function """
# 加了这些注释后，别人import你的module & class & function时 会显示这些详情提示

# ----------10_10 Pydoc 查看一个module的doc详细介绍
# 命令行：python3 -m pydoc math    |    python3 -m pydoc brightpypi.pdf2text
# python3 -m pydoc -p 1234 起一个本地server查看所有module


# ----------11_1-6 Popular python packages - Rest API
import requests
import config

url = 'https://restapi.amap.com/v3/weather/weatherInfo'

params = {
    # 'key': config.api_key,
    'key': '5ce7e6560f916d55c6a9a6ca965e9625',
    'city': '110101'
}
headers = {
    'Authorization': "Bearer " + config.api_key
}

response = requests.get(url=url, headers=headers, params=params)
weather_data = response.json()['lives'][0]
print(weather_data)


# ----------11_07 Selenium Webbrowser automation
# pip3 install selenium
# 去pypi搜selenium下载chromedriver，然后把它放到 /usr/local/bin下面
from selenium import webdriver
from selenium.webdriver.common.by import By

driver = webdriver.Chrome()
driver.get('https://github.com')

signin_link = driver.find_element(By.LINK_TEXT, "Sign in")
signin_link.click()

username_box = driver.find_element(By.ID, 'login_field')
username_box.send_keys('nzsyl0203@gmail.com')
password_box = driver.find_element(By.ID, 'password')
password_box.send_keys('Ying0203!github')
password_box.submit()

driver.quit()


# ----------11_08 Web Scraping/Crawler 网路爬虫
# pip3 install beautifulsoup4
import requests
from bs4 import BeautifulSoup

html = requests.get('https://stackoverflow.com/questions').text
soup = BeautifulSoup(html, 'html.parser')
questions = soup.select('.s-post-summary.js-post-summary')
for question in questions:
    print(question.select_one('.s-link').getText())
    print(question.select_one('.s-post-summary--stats-item-number').getText())
# titles = soup.find_all('h3')
# for title in titles:
#     print(title.text)


# ----------11_09 Working with PDF
# pip3 install pypdf2
import PyPDF2

with open('pdf/dummy.pdf', 'rb') as file:
    reader = PyPDF2.PdfReader(file)
    page = reader.pages[0]
    page.rotate(180)
    writer = PyPDF2.PdfWriter()
    writer.add_page(page)
    with open('pdf/rotated.pdf', 'wb') as output:
        writer.write(output)

merger = PyPDF2.PdfMerger()
pdfs = ['pdf/dummy.pdf', 'pdf/test.pdf']
for pdf in pdfs:
    merger.append(pdf)
merger.write('pdf/combined.pdf')

def merge_pdf(*pdfs):
    merger = PyPDF2.PdfMerger()
    for pdf in pdfs:
        merger.append(pdf)
    merger.write('pdf/combined.pdf')

merge_pdf('pdf/dummy.pdf', 'pdf/test.pdf', 'pdf/rotated.pdf')


# ----------11_10 Working with Excel spreadsheet
import openpyxl

wb = openpyxl.load_workbook('excel/transactions.xlsx')
# print(wb.sheetnames)    # ['Sheet1']
sheet = wb['Sheet1']
# print(sheet.max_row, sheet.max_column)

cell = sheet['a1']          # == sheet.cell(row=1, column=1)
# print(cell.value)
# print(cell.row)
# print(cell.column_letter)
# print(cell.coordinate)

cells_row = sheet[4]        # tuple of 4th row data => 1003  3  $7.95
cells_col = sheet['c']      # tuple of column C data => price $5.95 $6.95 $7.95
all_rows = sheet[1:4]       # tuple of tuples with all rows
all_columns = sheet['a:c']  # tuple of tuples with all columns
cells_range = sheet['a1:c3']

# for column in all_columns:
#     for cell in column:
#         print(cell.value)

# for column in range(1, sheet.max_column + 1):
#     for row in range(1, sheet.max_row + 1):
#         print(sheet.cell(column, row).value)

# for row in range(1, sheet.max_row + 1):
#     for column in range(1, sheet.max_column + 1):
#         cell = sheet.cell(row, column)
#         print(cell.value)

# wb2 = openpyxl.Workbook()
# wb2.create_sheet('Sheet2', 0)
# wb2.remove(wb2['sheet2'])

sheet.append([1, 2, 3])    # add a row at the end
# sheet.insert_rows() sheet.insert_cols() sheet.delete_cols() sheet.delete_rows()
wb.save('abc.xlsx')


# ----------11_11 Command Query Separation Principle
"""
命令查询分离原则（Command Query Separation Principle, CQS）
该原则由 Bertrand Meyer 提出，核心思想是：
- 命令（Command）：
·修改系统状态（如增删改数据），不返回结果（void 方法）
·例如：save(), delete()
- 查询（Query）：
·获取系统状态（如读取数据），返回结果（不修改状态）
·例如：getUser(), findAll()
-关键规则：
一个方法要么是<命令>，要么是<查询>，不要同时做两件事。
"""


# ----------11_12 NumPy
# pip install numpy
import numpy as np

array = np.array([[1, 2, 3], [4, 5, 6]])    # matrix
# print(array.shape)  => (2, 3), returns a tuple of 2 rows and 3 columns
zeros = np.zeros((3, 4), dtype=int)
ones = np.ones((3, 4), dtype=int)
fives = np.full((3, 4), 5)
randoms = np.random.random((3, 4))
matrix = np.random.rand(3, 4)
# print(randoms)
# print(randoms[1, 2])
# print(randoms > 0.5)    # returns a matrix of True or False
# print(randoms[randoms > 0.5])   # return a new array with value > 0.5
# print(np.sum(randoms))      # add all numbers
# np.ceil() np.floor() np.round()

# add minus mutiply division
first = np.array([1, 2, 3])
second = np.array([4, 5, 6])
# first + second = [5 7 9]; first + 2 = [3 4 5]; first * 2 = [2 4 6];


# ----------13_1-3 Machine Learning & libraries and tools
# 机器学习步骤：
# 1. import data and clean it, then split the data into training / test set
# 2. create a model and train it, then make predictions
# 3. evaluate and improve the accuracy and performance
# 工具集：
# 4. Libraries: Numpy, Pandas, MatPlotLib, Scikit-Learn...
# 5. tools: Jupyter, Anaconda
# 6. 关闭自动激活conda base环境：conda config --set auto_activate_base false (重开改为true)
# 7. 临时关闭激活：conda deactivate

# ----- 13_04 Jupyter notebook - load data
import pandas as pd
df = pd.read_csv('vgsales.csv')
df # df.shapes  df.describe() df.values

# ----- 13_05-08 Jupyter notebook - training and prediction
import pandas as pd
from sklearn.tree import DecisionTreeClassifier
# 加载数据
music_data = pd.read_csv('music.csv')
X = music_data.drop(columns=['genre'])
y = music_data['genre']
# 训练模型
model = DecisionTreeClassifier()
model.fit(X, y)
# 预测输出
test_data = pd.DataFrame([ [21, 1], [22, 0] ], columns=X.columns)
predictions = model.predict(test_data)
predictions

# ----- 13_09 Calculate the Accuracy by spliting data into training and test sets
import pandas as pd
from sklearn.tree import DecisionTreeClassifier
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
 
music_data = pd.read_csv('music.csv')
X = music_data.drop(columns=['genre'])
y = music_data['genre']
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2)

model = DecisionTreeClassifier()
model.fit(X_train, y_train)
predictions = model.predict(X_test)
score = accuracy_score(y_test, predictions)
score

# ----- 13_10 Persist and load trained model
# ---保存模型---
import pandas as pd
from sklearn.tree import DecisionTreeClassifier
import joblib

music_data = pd.read_csv('music.csv')
X = music_data.drop(columns=['genre'])
y = music_data['genre']

model = DecisionTreeClassifier()
model.fit(X, y)

joblib.dump(model, 'music_recom.joblib')

# ---加载模型---
import joblib
model = joblib.load('music_recom.joblib')
predictions = model.predict([ [21, 1], [23, 0] ])


# ----- 高阶函数（函数A把函数B和C作为参数）----- 匿名函数lambda
def calc_print(number, calculator, printer):
    result = calculator(number)
    printer(number, result)
    return result

def plus_ten(number):
    return number + 10

def times_five(number):
    return number * 5

def print_horizental(number, result):
    print(f'| Number: {number} | Result: {result} |')

def print_vertical(number, result):
    print(f'''
    | Number | {number} | 
    | Result | {result} |
    ''')

calc_print(30, plus_ten, print_vertical)
calc_print(10, times_five, print_horizental)
calc_print(5, lambda num: num**3, print_vertical)
print((lambda num1, num2: num1 + num2)(2, 3))


# ----- MVC 模型-视图-控制器 ----- 
# 模型 (Model) - 数据层
class PostModel:
    def __init__(self):
        self.posts = {
            1: {"id": 1, "title": "Python MVC", "content": "MVC模式解析"},
            2: {"id": 2, "title": "Django入门", "content": "Django基础教程"}
        }
    
    def get_post(self, post_id):
        """获取单篇文章"""
        return self.posts.get(post_id)
    
    def get_all_posts(self):
        """获取所有文章"""
        return list(self.posts.values())

# 视图 (View) - 展示层
class PostView:
    def render_post(self, post):
        """渲染单篇文章视图"""
        if not post:
            return "404 - 文章未找到"
        
        return f"""
        <html>
        <head><title>{post['title']}</title></head>
        <body>
            <h1>{post['title']}</h1>
            <p>{post['content']}</p>
        </body>
        </html>
        """
    
    def render_post_list(self, posts):
        """渲染文章列表视图"""
        html = "<html><body><h1>博客文章</h1><ul>"
        for post in posts:
            html += f"<li><a href='/posts/{post['id']}'>{post['title']}</a></li>"
        html += "</ul></body></html>"
        return html

# 控制器 (Controller) - 逻辑层
class PostController:
    def __init__(self):
        self.model = PostModel()
        self.view = PostView()
    
    def show_post(self, post_id):
        """显示单篇文章"""
        post = self.model.get_post(post_id)
        return self.view.render_post(post)
    
    def list_posts(self):
        """显示文章列表"""
        posts = self.model.get_all_posts()
        return self.view.render_post_list(posts)

# 使用示例
if __name__ == "__main__":
    controller = PostController()
    
    # 模拟请求：查看文章列表
    print("=== 文章列表 ===")
    print(controller.list_posts())
    
    # 模拟请求：查看单篇文章
    print("\n=== 单篇文章 ===")
    print(controller.show_post(1))